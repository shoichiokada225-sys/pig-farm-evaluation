# -*- coding: utf-8 -*-
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

JSON_PATH = r"C:\Users\so\pig-farm-evaluation\criteria-data.json"
OUT_PATH = r"C:\Users\so\Desktop\HSS実技試験_評価基準表_20260630.xlsx"

with open(JSON_PATH, encoding="utf-8") as f:
    data = json.load(f)

# ---- styles ----
HEADER_BG = "A8472A"
LEVEL_FILLS = {
    "level1": "FDE7E7",
    "level2": "FBEFE0",
    "level3": "FFF7E0",
    "level4": "EDF4E6",
    "level5": "E2EFE6",
}
thin = Side(style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)

HEADERS = ["No", "評価項目", "着眼点", "1 要指導", "2 要改善", "3 標準", "4 上回る", "5 模範的"]
COL_WIDTHS = [6, 18, 30, 36, 36, 36, 36, 36]
LEVEL_KEYS = ["level1", "level2", "level3", "level4", "level5"]


def style_header_row(ws, row):
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def set_col_widths(ws):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def page_setup(ws, header_row=1):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"


def est_row_height(texts, col_chars):
    """Estimate row height (points) from longest wrapped column."""
    max_lines = 1
    for txt, chars in zip(texts, col_chars):
        if not txt:
            continue
        lines = 0
        for seg in str(txt).split("\n"):
            # Japanese full-width: ~chars per line
            seg_len = len(seg)
            lines += max(1, -(-seg_len // chars))  # ceil
        max_lines = max(max_lines, lines)
    return max(110, min(max_lines * 15 + 8, 409))


wb = openpyxl.Workbook()

# ================= Sheet 1: レベルの考え方 =================
ws0 = wb.active
ws0.title = "レベルの考え方"
ws0.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

title_font = Font(bold=True, size=18, color="A8472A")
sub_font = Font(size=11, color="333333")
sec_font = Font(bold=True, size=13, color="A8472A")

ws0["A1"] = "HSS実技試験 評価基準表"
ws0["A1"].font = title_font
ws0["A2"] = "作成日: 2026-06-30"
ws0["A2"].font = sub_font

# scale table headers
SCALE_HEADERS = ["点数", "区分", "到達度の目安", "説明"]
SCALE_WIDTHS = [6, 14, 14, 80]
from openpyxl.utils import get_column_letter
for i, w in enumerate(SCALE_WIDTHS, start=1):
    ws0.column_dimensions[get_column_letter(i)].width = w

gyomu = [
    ("1", "要指導", "～30%", "大半の業務に指示・指導が必要。放置・見落とし・危険/不正確な手技。"),
    ("2", "要改善", "30%～", "指示・指摘があれば半数程度はできるが見落とし/抜けが多く、判断は人任せ。"),
    ("3", "標準", "50%～", "指示があれば自分で一通り作業でき、目に見える異常・事項は確認できる。"),
    ("4", "上回る", "70%～", "指示がなくても日常的に実施し、発熱を疑えば体温を実測する等、原因を探る確認行動・適切な初期対応ができる。"),
    ("5", "模範的", "90%～", "業務の意義を正しい知識で理解して判断・微調整し、記録・改善・新人指導までできる。"),
]
shisei = [
    ("1", "要指導", "～30%", "大半について都度確認・指示が必要。"),
    ("2", "要改善", "30%～", "時折抜け漏れがあるが、指摘・注意すれば行える。"),
    ("3", "標準", "50%～", "言われたことについては一通り実行できる。"),
    ("4", "上回る", "70%～", "発言や姿勢から、意識して実行しているのがわかる。"),
    ("5", "模範的", "90%～", "チーム・部署に働きかけができる。"),
]


def write_scale_table(ws, start_row, title, rows):
    ws.cell(row=start_row, column=1, value=title).font = sec_font
    hr = start_row + 1
    for c, h in enumerate(SCALE_HEADERS, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[hr].height = 22
    for ri, row in enumerate(rows):
        r = hr + 1 + ri
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            if c == 4:
                cell.alignment = BODY_ALIGN
            else:
                cell.alignment = CENTER_TOP
        ws.row_dimensions[r].height = est_row_height([row[3]], [70])
    return hr + 1 + len(rows)


next_row = write_scale_table(ws0, 4, "■ 業務（作業遂行能力）スケール", gyomu)
next_row = write_scale_table(ws0, next_row + 1, "■ 姿勢スケール", shisei)

# usage notes
note_row = next_row + 1
ws0.cell(row=note_row, column=1, value="■ 使い方").font = sec_font
notes = [
    "・各舎シートの評価項目を、それぞれ1〜5で採点してください。",
    "・採点に迷ったら、その項目の「着眼点」と各レベルの説明文を読み比べて、最も近いものを選びます。",
    "・レベルは下位の要件を満たした上で上位に到達している前提です（例：4は3の内容ができている）。",
    "・「仕事への姿勢（全舎共通）」シートは、どの舎の評価でも共通で使用します。",
    "・業務シートは作業遂行能力スケール、姿勢シートは姿勢スケールで採点します。",
]
for i, n in enumerate(notes):
    r = note_row + 1 + i
    cell = ws0.cell(row=r, column=1, value=n)
    cell.font = sub_font
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

ws0.freeze_panes = "A4"
ws0.page_setup.orientation = "landscape"
ws0.page_setup.paperSize = 8
ws0.page_setup.fitToWidth = 1
ws0.page_setup.fitToHeight = 0

# ================= Data sheets =================
col_chars = [4, 12, 22, 24, 24, 24, 24, 24]  # approx chars per line per column

for block in data:
    ws = wb.create_sheet(title=block["sheet"][:31])
    set_col_widths(ws)
    # header
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1)
    # body
    for ri, item in enumerate(block["items"]):
        r = ri + 2
        vals = [item["no"], item["name"], item["kanten"]] + [item[k] for k in LEVEL_KEYS]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            if c == 1:
                cell.alignment = CENTER_TOP
            else:
                cell.alignment = BODY_ALIGN
            if c >= 4:
                key = LEVEL_KEYS[c - 4]
                cell.fill = PatternFill("solid", fgColor=LEVEL_FILLS[key])
        ws.row_dimensions[r].height = est_row_height(vals, col_chars)
    ws.freeze_panes = "D2"
    page_setup(ws, header_row=1)

wb.save(OUT_PATH)
print("SAVED:", OUT_PATH)

# ================= Verify =================
wb2 = openpyxl.load_workbook(OUT_PATH)
print("SHEETS:", len(wb2.sheetnames))
expected = {b["sheet"][:31]: len(b["items"]) for b in data}
all_ok = True
print("--- data sheets ---")
for b in data:
    name = b["sheet"][:31]
    ws = wb2[name]
    data_rows = ws.max_row - 1  # minus header
    ok = (data_rows == len(b["items"]))
    all_ok = all_ok and ok
    print(f"{name}: rows={data_rows} expected={len(b['items'])} {'OK' if ok else 'NG'}")
print("first sheet:", wb2.sheetnames[0])
print("RESULT:", "OK" if (all_ok and wb2.sheetnames[0] == "レベルの考え方" and len(wb2.sheetnames) == len(data) + 1) else "NG")
